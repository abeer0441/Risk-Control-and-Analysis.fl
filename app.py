from flask import Flask,jsonify,request,render_template_string,Response,session,redirect,render_template,url_for
import pandas as pd
from openpyxl import load_workbook
import warnings,json
from flask_cors import CORS
from matplotlib.backends.backend_agg import FigureCanvasAgg as FigureCanvas
import matplotlib
matplotlib.use('agg')
import matplotlib.pyplot as plt
import seaborn as sn

#config = {
    #"origin":["http://localhost:3000"],
    #"methods":["OPTIONS","GET","POST"]
#}

app = Flask(__name__)
app.secret_key = b'_5#y2L"F4Q8z\n\xec]/'

CORS(app,resources={
    "/analyze": {"origins": "https://risk-analysis-and-control-b80708ff8ede.herokuapp.com/", "methods": ["OPTIONS", "GET", "POST"]},
    "/streamlit":{"origins": "https://risk-analysis-and-control-b80708ff8ede.herokuapp.com", "methods": ["OPTIONS", "GET", "POST"]},
    "/get_data": {"origin": "https://risk-analysis-and-control-b80708ff8ede.herokuapp.com" , "methods":["OPTIONS","GET","POST"]}
})

def analyze_excel_file(file):
    # Load the Excel file into a DataFrame
    warnings.simplefilter(action='ignore', category=UserWarning)
    wb = load_workbook(file)
    
    sheet = wb['Sheet1']

    information = {
        "sector_name": sheet.cell(row=3, column=7).value,
        "business_unit": sheet.cell(row=3, column=11).value,
        "project_cost": sheet.cell(row=3, column=15).value,
        "project_name": sheet.cell(row=5, column=7).value,
        "contractor": sheet.cell(row=5, column=11).value,
        "project_duration": sheet.cell(row=5, column=15).value,
    }
    return information

@app.route("/analyze",methods = ['POST'])
def analyze():
    if request.method == 'POST':
        if request.data is not None:
                f = request.files["file"]
                info = analyze_excel_file(f)
                df = pd.read_excel(f, skiprows=7, engine='openpyxl')
                data = {
                    "info":info,
                    "df":df.to_json()
                }
                session["data"] = data
                return jsonify(data)
            
        else:
            return jsonify([])


@app.route("/streamlit",methods=["POST"])
def streamlit():
    if request.data is not None:
        with open("data.json", 'w') as file:
            json.dump(request.get_json(), file)
    return jsonify({"status_code":200})

@app.route("/get_data",methods=["POST"])
def get_data():
    if request.data is not None:
        
            with open("data.json", 'r') as file:
                data = file.read()
                data = json.loads(data)
                df = pd.read_json(data["df"])
            return jsonify(data)

if __name__ == '__main__':
    app.run(debug=True)          